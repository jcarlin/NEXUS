"""Format helpers for court-formatted privilege log exports (CSV and XLSX)."""

from __future__ import annotations

import csv
import io

import openpyxl
from openpyxl.styles import Font

# Standard court-format column headers
_HEADERS = [
    "Bates Number",
    "Doc Date",
    "Author",
    "Recipient(s)",
    "Doc Type",
    "Subject",
    "Privilege Claimed",
    "Basis",
]


def format_privilege_log_csv(entries: list[dict]) -> io.StringIO:
    """Format privilege log entries as a CSV StringIO.

    Each entry dict should have keys matching ``PrivilegeLogEntry`` fields:
    bates_number, doc_date, author, recipients, doc_type, subject,
    privilege_claimed, basis.
    """
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL)
    writer.writerow(_HEADERS)
    for entry in entries:
        writer.writerow(
            [
                entry.get("bates_number", ""),
                entry.get("doc_date", ""),
                entry.get("author", ""),
                entry.get("recipients", ""),
                entry.get("doc_type", ""),
                entry.get("subject", ""),
                entry.get("privilege_claimed", ""),
                entry.get("basis", ""),
            ]
        )
    buf.seek(0)
    return buf


def format_privilege_log_xlsx(entries: list[dict]) -> io.BytesIO:
    """Format privilege log entries as an XLSX BytesIO.

    Same entry dict format as :func:`format_privilege_log_csv`.
    Header row is bold.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Privilege Log"

    bold = Font(bold=True)
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, entry in enumerate(entries, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("bates_number", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("doc_date", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("author", ""))
        ws.cell(row=row_idx, column=4, value=entry.get("recipients", ""))
        ws.cell(row=row_idx, column=5, value=entry.get("doc_type", ""))
        ws.cell(row=row_idx, column=6, value=entry.get("subject", ""))
        ws.cell(row=row_idx, column=7, value=entry.get("privilege_claimed", ""))
        ws.cell(row=row_idx, column=8, value=entry.get("basis", ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
