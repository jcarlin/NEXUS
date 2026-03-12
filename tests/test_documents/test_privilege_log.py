"""Tests for privilege log generation, export, and management."""

from __future__ import annotations

import csv
import io
from datetime import UTC, datetime
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import UUID, uuid4

import pytest
from httpx import AsyncClient

from app.documents.privilege_export import format_privilege_log_csv, format_privilege_log_xlsx
from app.documents.schemas import (
    PrivilegeBasisUpdate,
    PrivilegeLogEntry,
    PrivilegeLogExportFormat,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_MATTER_ID = UUID("00000000-0000-0000-0000-000000000001")


def _sample_entries() -> list[dict]:
    """Return sample privilege log entry dicts."""
    return [
        {
            "bates_number": "NEXUS000001 - NEXUS000005",
            "doc_date": "2024-06-15",
            "author": "Jane Smith",
            "recipients": "John Doe; Bob Jones",
            "doc_type": "correspondence",
            "subject": "Legal Strategy Discussion",
            "privilege_claimed": "Attorney-Client Privilege",
            "basis": "Communication between attorney and client regarding litigation strategy",
        },
        {
            "bates_number": "NEXUS000006",
            "doc_date": "2024-07-01",
            "author": "External Counsel",
            "recipients": "In-House Legal",
            "doc_type": "legal_filing",
            "subject": "Draft Motion Analysis",
            "privilege_claimed": "Work Product Doctrine",
            "basis": "Attorney mental impressions and legal analysis prepared in anticipation of litigation",
        },
    ]


# ---------------------------------------------------------------------------
# CSV format tests
# ---------------------------------------------------------------------------


class TestFormatCsv:
    def test_csv_with_sample_data(self) -> None:
        """CSV output has header + data rows."""
        entries = _sample_entries()
        buf = format_privilege_log_csv(entries)
        content = buf.getvalue()
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        assert len(rows) == 3  # header + 2 data rows
        assert rows[0][0] == "Bates Number"
        assert rows[0][-1] == "Basis"
        assert rows[1][0] == "NEXUS000001 - NEXUS000005"
        assert rows[1][2] == "Jane Smith"
        assert rows[2][4] == "legal_filing"

    def test_csv_empty_list(self) -> None:
        """CSV with no entries produces header-only output."""
        buf = format_privilege_log_csv([])
        content = buf.getvalue()
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        assert len(rows) == 1  # header only
        assert rows[0][0] == "Bates Number"

    def test_csv_special_characters(self) -> None:
        """CSV properly escapes commas, quotes, and newlines."""
        entries = [
            {
                "bates_number": "NEXUS000001",
                "doc_date": "2024-01-01",
                "author": 'O\'Brien, "Pat"',
                "recipients": "Smith, Jones",
                "doc_type": "email",
                "subject": "Re: Budget, Q4",
                "privilege_claimed": "Attorney-Client Privilege",
                "basis": "Privileged communication",
            },
        ]
        buf = format_privilege_log_csv(entries)
        content = buf.getvalue()

        # Re-parse to verify round-trip
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        assert len(rows) == 2
        assert rows[1][2] == 'O\'Brien, "Pat"'
        assert rows[1][3] == "Smith, Jones"


# ---------------------------------------------------------------------------
# XLSX format tests
# ---------------------------------------------------------------------------


class TestFormatXlsx:
    def test_xlsx_with_sample_data(self) -> None:
        """XLSX has header row with bold formatting and data rows."""
        entries = _sample_entries()
        buf = format_privilege_log_xlsx(entries)

        import openpyxl

        wb = openpyxl.load_workbook(buf)
        ws = wb.active

        assert ws.title == "Privilege Log"
        assert ws.cell(row=1, column=1).value == "Bates Number"
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.cell(row=1, column=8).value == "Basis"
        assert ws.cell(row=2, column=1).value == "NEXUS000001 - NEXUS000005"
        assert ws.cell(row=2, column=3).value == "Jane Smith"
        assert ws.cell(row=3, column=5).value == "legal_filing"
        assert ws.max_row == 3  # header + 2 data rows

    def test_xlsx_empty_list(self) -> None:
        """XLSX with no entries produces header-only worksheet."""
        buf = format_privilege_log_xlsx([])

        import openpyxl

        wb = openpyxl.load_workbook(buf)
        ws = wb.active

        assert ws.cell(row=1, column=1).value == "Bates Number"
        assert ws.max_row == 1  # header only


# ---------------------------------------------------------------------------
# Service method tests
# ---------------------------------------------------------------------------


class TestServiceGetEntries:
    @pytest.mark.asyncio
    async def test_returns_correct_entries(self) -> None:
        """get_privilege_log_entries returns formatted entries."""
        mock_db = AsyncMock()
        mock_row_1 = MagicMock()
        mock_row_1._mapping = {
            "id": uuid4(),
            "filename": "memo.pdf",
            "document_type": "correspondence",
            "created_at": datetime(2024, 6, 15, tzinfo=UTC),
            "privilege_status": "privileged",
            "privilege_basis": "Attorney-client communication",
            "bates_begin": "NEXUS000001",
            "bates_end": "NEXUS000005",
            "metadata_": {"author": "Jane Smith", "to": "John Doe"},
            "privilege_log_excluded": False,
        }
        mock_result = MagicMock()
        mock_result.all.return_value = [mock_row_1]
        mock_db.execute.return_value = mock_result

        from app.documents.service import DocumentService

        entries = await DocumentService.get_privilege_log_entries(
            db=mock_db,
            matter_id=_MATTER_ID,
        )

        assert len(entries) == 1
        assert entries[0]["privilege_claimed"] == "Attorney-Client Privilege"
        assert entries[0]["bates_number"] == "NEXUS000001 - NEXUS000005"
        assert entries[0]["author"] == "Jane Smith"
        assert entries[0]["recipients"] == "John Doe"
        assert entries[0]["basis"] == "Attorney-client communication"

    @pytest.mark.asyncio
    async def test_filters_by_matter_id(self) -> None:
        """get_privilege_log_entries includes matter_id in query params."""
        mock_db = AsyncMock()
        mock_result = MagicMock()
        mock_result.all.return_value = []
        mock_db.execute.return_value = mock_result

        from app.documents.service import DocumentService

        await DocumentService.get_privilege_log_entries(db=mock_db, matter_id=_MATTER_ID)

        call_args = mock_db.execute.call_args
        params = call_args[0][1]
        assert params["matter_id"] == _MATTER_ID

    @pytest.mark.asyncio
    async def test_excludes_excluded_by_default(self) -> None:
        """get_privilege_log_entries excludes privilege_log_excluded docs by default."""
        mock_db = AsyncMock()
        mock_result = MagicMock()
        mock_result.all.return_value = []
        mock_db.execute.return_value = mock_result

        from app.documents.service import DocumentService

        await DocumentService.get_privilege_log_entries(db=mock_db, matter_id=_MATTER_ID)

        call_args = mock_db.execute.call_args
        query_text = str(call_args[0][0])
        assert "privilege_log_excluded" in query_text

    @pytest.mark.asyncio
    async def test_includes_excluded_when_flag_set(self) -> None:
        """get_privilege_log_entries includes excluded docs when include_excluded=True."""
        mock_db = AsyncMock()
        mock_result = MagicMock()
        mock_result.all.return_value = []
        mock_db.execute.return_value = mock_result

        from app.documents.service import DocumentService

        await DocumentService.get_privilege_log_entries(
            db=mock_db,
            matter_id=_MATTER_ID,
            include_excluded=True,
        )

        call_args = mock_db.execute.call_args
        query_text = str(call_args[0][0])
        # WHERE clause should not filter on privilege_log_excluded
        # (it still appears in SELECT, but not as a filter condition)
        where_part = query_text.split("WHERE", 1)[1] if "WHERE" in query_text else ""
        assert "privilege_log_excluded = FALSE" not in where_part
        assert "privilege_log_excluded IS NULL" not in where_part


class TestServiceUpdateBasis:
    @pytest.mark.asyncio
    async def test_updates_fields(self) -> None:
        """update_privilege_basis returns updated dict."""
        doc_id = uuid4()
        mock_db = AsyncMock()
        mock_row = MagicMock()
        mock_row._mapping = {
            "id": doc_id,
            "privilege_basis": "Attorney-client communication",
            "privilege_log_excluded": False,
        }
        mock_result = MagicMock()
        mock_result.first.return_value = mock_row
        mock_db.execute.return_value = mock_result

        from app.documents.service import DocumentService

        result = await DocumentService.update_privilege_basis(
            db=mock_db,
            document_id=doc_id,
            matter_id=_MATTER_ID,
            basis="Attorney-client communication",
            excluded=False,
        )

        assert result is not None
        assert result["id"] == doc_id
        assert result["privilege_basis"] == "Attorney-client communication"
        assert result["privilege_log_excluded"] is False

    @pytest.mark.asyncio
    async def test_enforces_matter_scoping(self) -> None:
        """update_privilege_basis includes matter_id in WHERE clause."""
        mock_db = AsyncMock()
        mock_result = MagicMock()
        mock_result.first.return_value = None
        mock_db.execute.return_value = mock_result

        from app.documents.service import DocumentService

        result = await DocumentService.update_privilege_basis(
            db=mock_db,
            document_id=uuid4(),
            matter_id=_MATTER_ID,
            basis="test",
            excluded=False,
        )

        assert result is None
        call_args = mock_db.execute.call_args
        query_text = str(call_args[0][0])
        assert "matter_id" in query_text


# ---------------------------------------------------------------------------
# Router / endpoint tests
# ---------------------------------------------------------------------------


class TestPrivilegeLogEndpoint:
    @pytest.mark.asyncio
    async def test_get_privilege_log_returns_json(self, client: AsyncClient) -> None:
        """GET /privilege-log returns PrivilegeLogResponse."""
        entries = _sample_entries()
        with patch(
            "app.documents.service.DocumentService.get_privilege_log_entries",
            new_callable=AsyncMock,
            return_value=entries,
        ):
            response = await client.get("/api/v1/privilege-log")

        assert response.status_code == 200
        body = response.json()
        assert body["total"] == 2
        assert len(body["entries"]) == 2
        assert body["entries"][0]["bates_number"] == "NEXUS000001 - NEXUS000005"
        assert "matter_id" in body

    @pytest.mark.asyncio
    async def test_get_privilege_log_empty_matter(self, client: AsyncClient) -> None:
        """GET /privilege-log for empty matter returns empty list."""
        with patch(
            "app.documents.service.DocumentService.get_privilege_log_entries",
            new_callable=AsyncMock,
            return_value=[],
        ):
            response = await client.get("/api/v1/privilege-log")

        assert response.status_code == 200
        body = response.json()
        assert body["total"] == 0
        assert body["entries"] == []


class TestPrivilegeLogExportEndpoint:
    @pytest.mark.asyncio
    async def test_export_csv(self, client: AsyncClient) -> None:
        """GET /privilege-log/export?format=csv returns CSV content."""
        entries = _sample_entries()
        with patch(
            "app.documents.service.DocumentService.get_privilege_log_entries",
            new_callable=AsyncMock,
            return_value=entries,
        ):
            response = await client.get("/api/v1/privilege-log/export?format=csv")

        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "privilege_log.csv" in response.headers["content-disposition"]
        content = response.text
        assert "Bates Number" in content
        assert "NEXUS000001 - NEXUS000005" in content

    @pytest.mark.asyncio
    async def test_export_xlsx(self, client: AsyncClient) -> None:
        """GET /privilege-log/export?format=xlsx returns XLSX content."""
        entries = _sample_entries()
        with patch(
            "app.documents.service.DocumentService.get_privilege_log_entries",
            new_callable=AsyncMock,
            return_value=entries,
        ):
            response = await client.get("/api/v1/privilege-log/export?format=xlsx")

        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
        assert "privilege_log.xlsx" in response.headers["content-disposition"]
        # Verify it's valid XLSX by parsing it
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Bates Number"
        assert ws.max_row == 3

    @pytest.mark.asyncio
    async def test_export_without_format_returns_422(self, client: AsyncClient) -> None:
        """GET /privilege-log/export without format query param returns 422."""
        response = await client.get("/api/v1/privilege-log/export")
        assert response.status_code == 422


class TestPrivilegeBasisEndpoint:
    @pytest.mark.asyncio
    async def test_patch_privilege_basis(self, client: AsyncClient) -> None:
        """PATCH /documents/{id}/privilege-basis updates basis."""
        doc_id = uuid4()
        with patch(
            "app.documents.service.DocumentService.update_privilege_basis",
            new_callable=AsyncMock,
            return_value={
                "id": doc_id,
                "privilege_basis": "Attorney-client communication",
                "privilege_log_excluded": False,
            },
        ):
            response = await client.patch(
                f"/api/v1/documents/{doc_id}/privilege-basis",
                json={"privilege_basis": "Attorney-client communication", "privilege_log_excluded": False},
            )

        assert response.status_code == 200
        body = response.json()
        assert body["privilege_basis"] == "Attorney-client communication"
        assert body["privilege_log_excluded"] is False

    @pytest.mark.asyncio
    async def test_patch_privilege_basis_not_found(self, client: AsyncClient) -> None:
        """PATCH /documents/{id}/privilege-basis returns 404 for wrong matter."""
        with patch(
            "app.documents.service.DocumentService.update_privilege_basis",
            new_callable=AsyncMock,
            return_value=None,
        ):
            response = await client.patch(
                f"/api/v1/documents/{uuid4()}/privilege-basis",
                json={"privilege_basis": "test", "privilege_log_excluded": False},
            )

        assert response.status_code == 404


# ---------------------------------------------------------------------------
# Schema validation tests
# ---------------------------------------------------------------------------


class TestSchemaValidation:
    def test_privilege_log_entry_defaults(self) -> None:
        """PrivilegeLogEntry accepts all-None fields."""
        entry = PrivilegeLogEntry()
        assert entry.bates_number is None
        assert entry.doc_date is None
        assert entry.privilege_claimed is None

    def test_privilege_log_entry_with_data(self) -> None:
        """PrivilegeLogEntry accepts full data."""
        entry = PrivilegeLogEntry(
            bates_number="NEXUS000001",
            doc_date="2024-06-15",
            author="Jane Smith",
            recipients="John Doe",
            doc_type="email",
            subject="Test",
            privilege_claimed="Attorney-Client Privilege",
            basis="Privileged",
        )
        assert entry.bates_number == "NEXUS000001"
        assert entry.author == "Jane Smith"

    def test_privilege_log_export_format_values(self) -> None:
        """PrivilegeLogExportFormat has csv and xlsx values."""
        assert PrivilegeLogExportFormat.CSV == "csv"
        assert PrivilegeLogExportFormat.XLSX == "xlsx"
        assert len(PrivilegeLogExportFormat) == 2

    def test_privilege_basis_update_defaults(self) -> None:
        """PrivilegeBasisUpdate has correct defaults."""
        update = PrivilegeBasisUpdate()
        assert update.privilege_basis is None
        assert update.privilege_log_excluded is False

    def test_privilege_basis_update_with_data(self) -> None:
        """PrivilegeBasisUpdate accepts explicit values."""
        update = PrivilegeBasisUpdate(
            privilege_basis="Attorney-client communication",
            privilege_log_excluded=True,
        )
        assert update.privilege_basis == "Attorney-client communication"
        assert update.privilege_log_excluded is True
